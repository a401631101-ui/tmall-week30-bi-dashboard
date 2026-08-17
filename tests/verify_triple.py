#!/usr/bin/env python3
"""
三端核对：Excel ↔ JSON（reports） ↔ 页面数据源（facts + index）
用法: python3 tests/verify_triple.py [周号 ...]   # 不传则核对全部 32 周
校验：
  1. 周号一致性 —— index / facts / reports 三处周号对齐（1..32）
  2. facts ↔ reports 一致性 —— reports 的 top10 商品ID ⊆ 同周 facts 商品ID
  3. Excel ↔ JSON 一致性 —— 首页总览 KPI 单元格 == reports.meta.quickMetrics
"""
import json
import os
import sys

import openpyxl

BASE = '/Users/fengna/Documents/Codex/天猫类目bi看板 tmall-week30-bi-dashboard /第30周天猫重点品类市场竞品周报/'
DATA = '/Users/fengna/Documents/Codex/天猫类目bi看板 tmall-week30-bi-dashboard /tmall-week30-bi-dashboard/public/data/'
XLSX_DIR = BASE + '竞品周报/'

failures = []


def check(cond, msg):
    if not cond:
        failures.append(msg)
        print(f'  ✗ {msg}')
    return cond


def load_json(path):
    with open(path, encoding='utf-8') as f:
        return json.load(f)


def main():
    weeks = [int(a) for a in sys.argv[1:]] or list(range(1, 33))
    print(f'核对周: {weeks}')

    index = load_json(DATA + 'index.json')
    index_weeks = {w['week'] for w in index}

    # 1. 周号一致性
    print('\n[1] 周号一致性 (index / facts / reports)')
    facts_files = sorted(os.listdir(DATA + 'facts'))
    reports_files = sorted(os.listdir(DATA + 'reports'))
    facts_weeks = {int(f[5:7]) for f in facts_files if f.startswith('week-') and f.endswith('.json')}
    reports_weeks = {int(f[5:7]) for f in reports_files if f.startswith('week-') and f.endswith('.json')}
    check(index_weeks == set(range(1, 33)), f'index 覆盖 1..32 (实际 {sorted(index_weeks)})')
    check(facts_weeks == set(range(1, 33)), f'facts 覆盖 1..32 (实际 {sorted(facts_weeks)})')
    check(reports_weeks == set(range(1, 33)), f'reports 覆盖 1..32 (实际 {sorted(reports_weeks)})')

    # 2. facts ↔ reports 一致性
    print('\n[2] facts ↔ reports 商品ID 一致性')
    for w in weeks:
        if w not in reports_weeks or w not in facts_weeks:
            continue
        facts = load_json(DATA + f'facts/week-{w:02d}.json')
        report = load_json(DATA + f'reports/week-{w:02d}.json')
        fact_ids = {r['productId2'] for r in facts if r['productId2']}
        # reports top10 商品ID 应能在同周 facts 中找到
        missing = []
        for p in report['top10Products']:
            if p['productId'] and p['productId'] not in fact_ids:
                missing.append(p['productId'])
        check(report['meta']['week'] == w, f'W{w} reports.meta.week == {w}')
        check(not missing, f'W{w} reports top10 有 {len(missing)} 个商品ID不在 facts 中: {missing[:5]}')
        # facts 内部周号自洽
        bad = [r['productId2'] for r in facts if r['week'] != w]
        check(not bad, f'W{w} facts 有 {len(bad)} 行 week 字段 != {w}')

    # 3. Excel ↔ JSON 一致性（首页总览 KPI）
    print('\n[3] Excel ↔ JSON (首页总览 KPI)')
    for w in weeks:
        xlsx = XLSX_DIR + f'第{w}周天猫重点品类市场竞品周报_Excel版.xlsx'
        if not os.path.exists(xlsx):
            check(False, f'W{w} Excel 文件缺失: {xlsx}')
            continue
        if w not in reports_weeks:
            continue
        report = load_json(DATA + f'reports/week-{w:02d}.json')
        qm = report['meta']['quickMetrics']
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        ws = wb['首页总览']
        # 模板布局: B20=t_cats D20=t_rows F20=t_top10 H20=anomaly B21=brand_count D21=src_total_rows
        excel = {
            'categoryCount': ws['B20'].value,
            'totalRows': ws['D20'].value,
            'top10Count': ws['F20'].value,
            'anomalyCount': ws['H20'].value,
            'brandCount': ws['B21'].value,
            'sourceTotalRows': ws['D21'].value,
        }
        wb.close()
        for key, expect in [('categoryCount', qm['categoryCount']), ('totalRows', qm['totalRows']),
                            ('top10Count', qm['top10Count']), ('anomalyCount', qm['anomalyCount']),
                            ('brandCount', qm['brandCount']), ('sourceTotalRows', qm['sourceTotalRows'])]:
            check(excel[key] == expect, f'W{w} Excel.{key}={excel[key]} vs JSON.{key}={expect}')

    # 汇总
    print('\n' + '=' * 50)
    if failures:
        print(f'✗ 核对失败 {len(failures)} 项')
        sys.exit(1)
    print(f'✓ 三端核对通过（{len(weeks)} 周）')


if __name__ == '__main__':
    main()
